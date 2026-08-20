"""Generate the expert-review workbook for a set of question runs.

The review format itself was arrived at by the reviewers and works: one tab
per question, the full ranked retrieval set above, the answer broken into
claims below, and a verdict block at the end. What did not work was building
it by hand. The sheet was assembled by pasting exported blocks into a
spreadsheet, and a paste that lands one column off silently reattributes
every verdict after it — which is exactly what happened, and cost a round of
review arguing about defects that were transcription, not output.

So the workbook is generated from the database instead, and the reviewers'
own vocabularies become real dropdowns (Excel data validation, which Google
Sheets imports as dropdown chips) rather than free text that has to be
retyped consistently.

Nothing here is specific to a deployment. The question list, the tab titles
and the workbook title come from a manifest file; the vocabularies are the
constants below.

    python -m app.review_export --manifest questions.yaml --out review.xlsx

Manifest shape (a list, or {"questions": [...]}):

    - id: "1"                     # tab is named Q01
      title: Market definition    # shown in the index
      scope: EU + UK              # free text, optional
      question: |                 # matched against the runs
        The relevant market in ...
"""

from __future__ import annotations

import argparse
import asyncio
import re
import uuid
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import text

from app.db import AsyncSessionLocal

# The reviewers' own words, taken from the verdicts they have already
# recorded. Changing these changes what a reviewer can choose, so they are
# deliberately in one place and deliberately short — a long dropdown is a
# free-text field with extra steps.
CLAIM_VERDICTS = ["Supports claim", "Partially supports claim",
                  "Does not support claim", "Wrong source",
                  "Missing citation", "Not a legal claim"]
REQUIRED_SOURCE = ["Yes", "No", "Partially", "Unknown"]
SOURCE_VERDICTS = ["Relevant", "Not relevant", "Duplicate",
                   "Missing important source"]
ACCEPTABLE = ["Yes", "No", "Partly"]
COMPLETED = ["Not started", "In review", "Review completed", "Needs follow-up"]

HEAD = PatternFill("solid", fgColor="EFEAE2")
SECTION = PatternFill("solid", fgColor="DCE5DC")
# Every cell the reviewer fills in is this colour and nothing else is, so
# "what is mine to do" is answerable by looking rather than by reading.
INPUT = PatternFill("solid", fgColor="FFF2CC")
BOLD = Font(bold=True)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")


def _tab(qid: str) -> str:
    """Q01, Q29 — zero-padded so tabs sort the way a reader expects."""
    digits = re.sub(r"\D", "", str(qid)) or "0"
    return f"Q{int(digits):02d}"


def _similar(a: str, b: str) -> float:
    ta = {w for w in (a or "").lower().split() if len(w) > 4}
    tb = {w for w in (b or "").lower().split() if len(w) > 4}
    return len(ta & tb) / len(ta | tb) if ta and tb else 0.0


async def _fetch(session, question: str) -> dict[str, Any] | None:
    """The most recent settled run for this question, with everything the
    workbook shows. Matching is on wording overlap: the manifest phrases a
    question in the reviewers' words and the run stores the wording that was
    actually asked, and the two are rarely identical."""
    runs = (await session.execute(text("""
        SELECT id, question, status, answer_id, parent_result_id, created_at
        FROM query_run
        WHERE status IN ('published', 'partial')
        ORDER BY created_at DESC"""))).mappings().all()
    best, score = None, 0.0
    for r in runs:
        s = _similar(question, r["question"])
        if s > score:
            best, score = r, s
    if not best or score < 0.45:
        return None

    claims = (await session.execute(text("""
        SELECT c.id, c.sequence, c.claim_text, c.claim_type, c.rationale
        FROM answer_claim c WHERE c.answer_id = :a ORDER BY c.sequence"""),
        {"a": best["answer_id"]})).mappings().all() if best["answer_id"] else []

    ev = (await session.execute(text("""
        SELECT e.claim_id, d.filename, e.span, e.validated,
               e.validation_reasoning, v.content_md
        FROM claim_evidence e
        JOIN answer_claim c ON c.id = e.claim_id
        JOIN document d ON d.id = e.document_id
        LEFT JOIN document_version v ON v.id = d.current_version_id
        WHERE c.answer_id = :a
        ORDER BY c.sequence"""),
        {"a": best["answer_id"]})).mappings().all() if best["answer_id"] else []

    docs = (await session.execute(text("""
        SELECT rd.rank, d.filename, dc.name AS class_name,
               -- property values are stored wrapped ({"_": "..."}); a
               -- scalar is possible too, so fall back rather than assume
               (SELECT coalesce(pv.value ->> '_', pv.value #>> '{}')
                FROM property_value pv
                JOIN document_class_property p ON p.id = pv.property_id
                WHERE pv.document_id = d.id AND p.name = 'title' LIMIT 1) AS title
        FROM result_document rd
        JOIN document d ON d.id = rd.document_id
        LEFT JOIN document_class dc ON dc.id = d.document_class_id
        WHERE rd.result_id = :r
        ORDER BY rd.rank"""),
        {"r": best["parent_result_id"]})).mappings().all() \
        if best["parent_result_id"] else []

    return {"run": dict(best), "claims": [dict(c) for c in claims],
            "evidence": [dict(e) for e in ev], "docs": [dict(d) for d in docs],
            "match": round(score, 2)}


def _passage(content: str | None, span: dict | None) -> str:
    """The lines the claim actually cites, as the reviewer will read them."""
    if not content or not span or span.get("line_from") is None:
        return ""
    lf = int(span["line_from"])
    lt = int(span.get("line_to") or lf)
    body = "\n".join(content.splitlines()[max(0, lf - 1):lt]).strip()
    return body[:1500]


def _dv(ws, options: list[str], cells: str) -> None:
    """A dropdown Google Sheets will also honour: an inline list, which
    survives the xlsx import as a dropdown chip. A cross-sheet range would
    not — it imports as a reference the sheet cannot resolve."""
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(cells)


def _write_question(wb: Workbook, entry: dict, data: dict | None) -> dict:
    ws = wb.create_sheet(_tab(entry.get("id", "0")))
    widths = [10, 46, 44, 26, 52, 24, 34, 16, 22, 40, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def section(title: str) -> int:
        ws.append([title])
        r = ws.max_row
        ws.cell(r, 1).font = BOLD
        for c in range(1, 12):
            ws.cell(r, c).fill = SECTION
        return r

    ws.append([f"{_tab(entry.get('id','0'))} — Expert review"])
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.append([])
    section("Question")
    for label, value in (("Question ID", str(entry.get("id", ""))),
                         ("Topic", entry.get("title", "")),
                         ("Scope", entry.get("scope", ""))):
        ws.append([label, value])
        ws.cell(ws.max_row, 1).font = BOLD
    ws.append(["Question", (entry.get("question") or "").strip()])
    ws.cell(ws.max_row, 1).font = BOLD
    ws.cell(ws.max_row, 2).alignment = WRAP

    if not data:
        ws.append([])
        ws.append(["No run found for this question."])
        return {"tab": _tab(entry.get("id", "0")), "status": "NOT RUN",
                "claims": 0, "verdict_row": None}

    run = data["run"]
    ws.append([])
    section("Generated answer")
    by_claim: dict[Any, list[dict]] = {}
    for e in data["evidence"]:
        by_claim.setdefault(e["claim_id"], []).append(e)
    ws.append(["", " ".join(c["claim_text"] for c in data["claims"])])
    ws.cell(ws.max_row, 2).alignment = WRAP
    ws.row_dimensions[ws.max_row].height = 90
    ws.append(["Run ID", str(run["id"]), "Run date",
               run["created_at"].strftime("%Y-%m-%d %H:%M") if run["created_at"] else "",
               "Outcome", run["status"]])
    ws.cell(ws.max_row, 1).font = BOLD

    # Reference numbers, so a citation can be found in the ranked list
    # below without searching for a filename. The number IS the rank.
    rank_of = {d["filename"]: d["rank"] for d in data["docs"]}

    def ref(fn: str) -> str:
        r = rank_of.get(fn)
        return f"[{r}] {fn}" if r else fn

    def numbered(prose: str) -> str:
        """Give any filename the reasoning mentions its reference number.

        The reasoning is told to name a source by its citation rather than by
        the file it is stored under, and mostly does. Where a filename gets
        through, a reader should at least be able to find it in the list
        below instead of meeting an opaque string.
        """
        out = prose or ""
        for fn in sorted(rank_of, key=len, reverse=True):
            if fn in out:
                out = out.replace(fn, ref(fn))
        return out

    # ── claim by claim ───────────────────────────────────────────────────
    # Claims come before the retrieval set: the answer is what a reviewer
    # reads first, and the ranked list runs to a hundred rows.
    ws.append([])
    section("Claim-by-claim review")
    ws.append(["Claim ID", "Claim from answer", "Reasoning behind the claim",
               "Cited source", "Cited passage", "Expert verdict",
               "Required fix / comment"])
    hdr = ws.max_row
    for c in range(1, 8):
        ws.cell(hdr, c).font = BOLD
        ws.cell(hdr, c).fill = HEAD
    for c in data["claims"]:
        rows = by_claim.get(c["id"], [])
        docs_ = "\n".join(sorted({ref(e["filename"]) for e in rows}))
        passages = "\n\n".join(
            f"{ref(e['filename'])} l.{(e['span'] or {}).get('line_from')}-"
            f"{(e['span'] or {}).get('line_to')}\n"
            f"{_passage(e['content_md'], e['span'])}" for e in rows)
        if c["claim_type"] == "abstention":
            docs_ = docs_ or "(no source — the answer states this is not established)"
        ws.append([c["sequence"], c["claim_text"],
                   numbered(c["rationale"] or ""), docs_, passages, "", ""])
        for col in (2, 3, 4, 5, 7):
            ws.cell(ws.max_row, col).alignment = WRAP
        for col in (6, 7):
            ws.cell(ws.max_row, col).fill = INPUT
        ws.row_dimensions[ws.max_row].height = 96
    first, last = hdr + 1, ws.max_row
    if last >= first:
        _dv(ws, CLAIM_VERDICTS, f"F{first}:F{last}")

    # ── retrieval set ────────────────────────────────────────────────────
    ws.append([])
    section("Retrieval set (full ranked list)")
    cited = {e["filename"] for e in data["evidence"]}
    ws.append(["Rank", "Title", "Document name", "Document class",
               "Cited in answer", "Required source?", "Expert verdict", "Notes"])
    hdr = ws.max_row
    for c in range(1, 9):
        ws.cell(hdr, c).font = BOLD
        ws.cell(hdr, c).fill = HEAD
    for d in data["docs"]:
        ws.append([d["rank"], d["title"] or d["filename"], d["filename"],
                   d["class_name"] or "",
                   "YES" if d["filename"] in cited else "", "", "", ""])
        for col in (6, 7, 8):
            ws.cell(ws.max_row, col).fill = INPUT
    first, last = hdr + 1, ws.max_row
    if last >= first:
        _dv(ws, REQUIRED_SOURCE, f"F{first}:F{last}")
        _dv(ws, SOURCE_VERDICTS, f"G{first}:G{last}")

    # ── verdict ──────────────────────────────────────────────────────────
    ws.append([])
    section("Final expert verdict")
    ws.append(["Reviewer", ""])
    ws.append(["Is the answer acceptable?", ""])
    acceptable_row = ws.max_row
    ws.append(["Overall comment / required fixes", ""])
    ws.append(["Review completed?", ""])
    completed_row = ws.max_row
    for r in range(acceptable_row - 1, completed_row + 1):
        ws.cell(r, 1).font = BOLD
        ws.cell(r, 2).fill = INPUT
    ws.cell(acceptable_row + 1, 2).alignment = WRAP
    _dv(ws, ACCEPTABLE, f"B{acceptable_row}")
    _dv(ws, COMPLETED, f"B{completed_row}")
    ws.freeze_panes = "A2"
    return {"tab": ws.title, "status": run["status"], "claims": len(data["claims"]),
            "reviewer_row": acceptable_row - 1, "acceptable_row": acceptable_row,
            "comment_row": acceptable_row + 1, "completed_row": completed_row}


def _write_index(wb: Workbook, entries: list[dict], meta: list[dict],
                 title: str) -> None:
    ws = wb.create_sheet("Review status", 0)
    for i, w in enumerate([12, 40, 18, 52, 20, 22, 18, 44], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.append([title])
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Question ID", "Topic", "Scope", "Question", "Review status",
               "Acceptable?", "Reviewer", "Main issue / follow-up"])
    for c in range(1, 9):
        ws.cell(ws.max_row, c).font = BOLD
        ws.cell(ws.max_row, c).fill = HEAD
    for entry, m in zip(entries, meta):
        # live references, so the index reflects each tab as it is filled in
        t = m["tab"]
        ref = (lambda r: f"='{t}'!B{r}") if m.get("acceptable_row") else (lambda r: "")
        ws.append([str(entry.get("id", "")), entry.get("title", ""),
                   entry.get("scope", ""), (entry.get("question") or "").strip(),
                   ref(m["completed_row"]) if m.get("completed_row") else m["status"],
                   ref(m["acceptable_row"]) if m.get("acceptable_row") else "",
                   ref(m["reviewer_row"]) if m.get("reviewer_row") else "",
                   ref(m["comment_row"]) if m.get("comment_row") else ""])
        ws.cell(ws.max_row, 4).alignment = WRAP
    ws.freeze_panes = "A4"


def _write_readme(wb: Workbook, title: str) -> None:
    ws = wb.create_sheet("Read me", 0)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 104
    ws.append([title])
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.append([])
    for label, body in [
        ("What this is", "One tab per question. Each tab holds the answer that "
                         "was generated, the documents retrieved for it, and "
                         "the answer broken into individual claims."),
        ("How to review", "Work down the claim-by-claim table, which comes "
                          "first. For each claim, read the cited passage — it "
                          "is the exact text the claim rests on — and set the "
                          "expert verdict."),
        ("Reasoning", "Each claim carries the reasoning behind it: which part "
                      "of the question the claim answers, why it cites the "
                      "source it does, and where it chose one source over "
                      "another. It is reasoning, not evidence — judge the claim "
                      "against the cited passage, and use this column to see "
                      "how the claim was arrived at."),
        ("Yellow cells", "Everything shaded yellow is yours to fill in. "
                         "Nothing else in the workbook is yellow."),
        ("Reference numbers", "A cited source is written [4] filename. The "
                              "number is its rank in the retrieval set below, "
                              "so you can find it without searching."),
        ("Comments", "Add a comment only where something needs to change. A "
                     "verdict with no comment reads as 'correct as written'."),
        ("Retrieval set", "The ranked list below the claims is every document "
                          "retrieved, whether or not the answer used it. Mark "
                          "anything that should have been used and was not."),
        ("Verdict", "Finish each tab with the block at the bottom. The review "
                    "status tab summarises those blocks automatically."),
        ("Dropdowns", "The verdict columns are dropdowns. If you open this in "
                      "Google Sheets, import it rather than pasting into an "
                      "existing sheet — a paste that lands one column off "
                      "silently reattributes every verdict after it."),
    ]:
        ws.append([label, body])
        ws.cell(ws.max_row, 1).font = BOLD
        ws.cell(ws.max_row, 2).alignment = WRAP
        ws.row_dimensions[ws.max_row].height = 46


def _write_lists(wb: Workbook) -> None:
    """The vocabularies, visible so a reviewer can see the full set at a
    glance. The dropdowns do not read from here — they carry their own
    inline list, which is what survives the Google Sheets import."""
    ws = wb.create_sheet("Lists")
    cols = [("Claim verdict", CLAIM_VERDICTS), ("Required source?", REQUIRED_SOURCE),
            ("Source verdict", SOURCE_VERDICTS), ("Acceptable?", ACCEPTABLE),
            ("Review status", COMPLETED)]
    for i, (head, values) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 28
        ws.cell(1, i, head).font = BOLD
        ws.cell(1, i).fill = HEAD
        for j, v in enumerate(values, start=2):
            ws.cell(j, i, v)


async def build(manifest: Path, out: Path, title: str) -> None:
    raw = yaml.safe_load(manifest.read_text())
    entries = raw.get("questions") if isinstance(raw, dict) else raw
    entries = [e for e in (entries or []) if e.get("question")]
    if not entries:
        raise SystemExit(f"no questions in {manifest}")

    wb = Workbook()
    wb.remove(wb.active)
    meta = []
    async with AsyncSessionLocal() as session:
        for e in entries:
            data = await _fetch(session, e["question"])
            meta.append(_write_question(wb, e, data))
            print(f"  {meta[-1]['tab']:5} {meta[-1]['status']:10} "
                  f"{meta[-1]['claims']:>3} claims", flush=True)
    _write_lists(wb)
    _write_index(wb, entries, meta, title)
    _write_readme(wb, title)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    missing = sum(1 for m in meta if m["status"] == "NOT RUN")
    print(f"\n{out}  ({len(entries)} questions"
          + (f", {missing} with no run" if missing else "") + ")")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--manifest", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    ap.add_argument("--title", default="Expert review")
    args = ap.parse_args()
    asyncio.run(build(args.manifest, args.out, args.title))


if __name__ == "__main__":
    main()
