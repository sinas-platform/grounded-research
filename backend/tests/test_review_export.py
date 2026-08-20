"""The review workbook's columns must line up with its data.

The format this replaces was assembled by pasting exported blocks into a
spreadsheet. A paste landing one column off silently reattributes every
verdict after it, which is not visible in the sheet and cost a round of
review arguing about defects that were transcription rather than output.

Generating it removes the paste. These tests pin the part that could still
drift: that each dropdown covers exactly the rows it is meant to judge, that
the index points at verdict cells that exist, and that a question with no run
still produces a tab rather than vanishing from the workbook.

Run from the backend directory: `python -m pytest tests/test_review_export.py`
"""

from __future__ import annotations

import uuid
from datetime import datetime

from openpyxl import Workbook

from app.review_export import (
    CLAIM_VERDICTS,
    _tab,
    _write_index,
    _write_question,
)

ENTRY = {"id": "Q3", "title": "Market boundary", "scope": "UK",
         "question": "Is there authority for treating grocery delivery "
                     "and restaurant delivery as separate markets?"}


def _data(n_claims: int = 3, n_docs: int = 5) -> dict:
    ids = [uuid.uuid4() for _ in range(n_claims)]
    return {
        "run": {"id": uuid.uuid4(), "question": ENTRY["question"],
                "status": "published", "answer_id": uuid.uuid4(),
                "parent_result_id": uuid.uuid4(),
                "created_at": datetime(2026, 8, 20, 15, 34)},
        "claims": [{"id": ids[i], "sequence": i + 1,
                    "claim_text": f"Claim number {i + 1} says a thing.",
                    "claim_type": "legal_principle",
                    "rationale": f"Why claim {i + 1} rests on this source."}
                   for i in range(n_claims)],
        "evidence": [{"claim_id": ids[i], "filename": f"doc{i}.md",
                      "span": {"line_from": 10, "line_to": 12},
                      "validated": True, "validation_reasoning": "supports",
                      "content_md": "\n".join(f"line {n}" for n in range(1, 30))}
                     for i in range(n_claims)],
        "docs": [{"rank": i + 1, "filename": f"doc{i}.md",
                  "class_name": "Regulatory Decision", "title": f"Document {i}"}
                 for i in range(n_docs)],
        "match": 0.9,
    }


def _ranges(ws) -> dict[str, str]:
    return {dv.formula1: str(dv.sqref) for dv in ws.data_validations.dataValidation}


def test_tab_names_sort_the_way_a_reader_expects():
    assert [_tab(q) for q in ("Q1", "3", "Q10", "Q29")] == ["Q01", "Q03", "Q10", "Q29"]


def test_each_dropdown_covers_exactly_its_own_rows():
    wb = Workbook()
    wb.remove(wb.active)
    _write_question(wb, ENTRY, _data(n_claims=3, n_docs=5))
    ws = wb["Q03"]

    claim_hdr = next(r for r in range(1, ws.max_row + 1)
                     if ws.cell(r, 1).value == "Claim ID")
    verdict_col = next(c for c in range(1, 9)
                       if ws.cell(claim_hdr, c).value == "Expert verdict")
    assert verdict_col == 6

    rng = _ranges(ws)['"' + ",".join(CLAIM_VERDICTS) + '"']
    assert rng == f"F{claim_hdr + 1}:F{claim_hdr + 3}"
    # and the rows it covers really are the claims, in order
    for offset in range(3):
        assert ws.cell(claim_hdr + 1 + offset, 1).value == offset + 1

    # the retrieval dropdowns cover the five document rows and stop there,
    # so a reviewer cannot mark a verdict on the section heading below
    src_hdr = next(r for r in range(1, ws.max_row + 1)
                   if ws.cell(r, 1).value == "Rank")
    ranges = set(_ranges(ws).values())
    assert f"F{src_hdr + 1}:F{src_hdr + 5}" in ranges, ranges
    assert f"G{src_hdr + 1}:G{src_hdr + 5}" in ranges, ranges
    assert ws.cell(src_hdr + 5, 3).value == "doc4.md"
    assert ws.cell(src_hdr + 6, 3).value is None


def test_a_claims_row_carries_its_passage_and_its_reasoning():
    wb = Workbook()
    wb.remove(wb.active)
    _write_question(wb, ENTRY, _data(n_claims=2, n_docs=2))
    ws = wb["Q03"]
    hdr = next(r for r in range(1, ws.max_row + 1)
               if ws.cell(r, 1).value == "Claim ID")
    headers = [ws.cell(hdr, c).value for c in range(1, 8)]
    assert headers == ["Claim ID", "Claim from answer",
                       "Reasoning behind the claim", "Cited source",
                       "Cited passage", "Expert verdict",
                       "Required fix / comment"]
    row = hdr + 1
    assert ws.cell(row, 3).value == "Why claim 1 rests on this source."
    # a citation carries its rank, so it can be found in the list below
    assert ws.cell(row, 4).value == "[1] doc0.md"
    assert ws.cell(row, 5).value.startswith("[1] doc0.md l.10-12")
    assert "line 10" in ws.cell(row, 5).value and "line 12" in ws.cell(row, 5).value


def test_claims_come_before_the_retrieval_set():
    """A reviewer reads the answer first; the ranked list runs to a hundred
    rows and used to sit on top of it."""
    wb = Workbook()
    wb.remove(wb.active)
    _write_question(wb, ENTRY, _data())
    ws = wb["Q03"]
    heads = [(r, ws.cell(r, 1).value) for r in range(1, ws.max_row + 1)]
    order = [v for _, v in heads
             if v in ("Claim-by-claim review", "Retrieval set (full ranked list)",
                      "Final expert verdict")]
    assert order == ["Claim-by-claim review", "Retrieval set (full ranked list)",
                     "Final expert verdict"]


def test_only_the_reviewers_own_cells_are_yellow():
    """"What is mine to fill in" should be answerable by looking."""
    from app.review_export import INPUT

    wb = Workbook()
    wb.remove(wb.active)
    meta = _write_question(wb, ENTRY, _data(n_claims=2, n_docs=3))
    ws = wb["Q03"]
    yellow = {(c.row, c.column) for row in ws.iter_rows() for c in row
              if c.fill is not None and c.fill.fgColor.rgb == INPUT.fgColor.rgb}

    claim_hdr = next(r for r in range(1, ws.max_row + 1)
                     if ws.cell(r, 1).value == "Claim ID")
    src_hdr = next(r for r in range(1, ws.max_row + 1)
                   if ws.cell(r, 1).value == "Rank")
    expected = {(claim_hdr + i, c) for i in (1, 2) for c in (6, 7)}
    expected |= {(src_hdr + i, c) for i in (1, 2, 3) for c in (6, 7, 8)}
    expected |= {(r, 2) for r in range(meta["reviewer_row"],
                                       meta["completed_row"] + 1)}
    assert yellow == expected, sorted(yellow ^ expected)


def test_the_index_points_at_verdict_cells_that_exist():
    wb = Workbook()
    wb.remove(wb.active)
    meta = [_write_question(wb, ENTRY, _data())]
    _write_index(wb, [ENTRY], meta, "Expert review")
    ws = wb["Review status"]
    formula = ws.cell(4, 6).value                      # "Acceptable?"
    assert formula == f"='Q03'!B{meta[0]['acceptable_row']}"
    assert wb["Q03"].cell(meta[0]["acceptable_row"], 1).value == \
        "Is the answer acceptable?"
    assert wb["Q03"].cell(meta[0]["completed_row"], 1).value == "Review completed?"
    assert wb["Q03"].cell(meta[0]["reviewer_row"], 1).value == "Reviewer"


def test_a_question_with_no_run_still_gets_a_tab():
    wb = Workbook()
    wb.remove(wb.active)
    meta = _write_question(wb, ENTRY, None)
    assert meta["status"] == "NOT RUN" and meta["claims"] == 0
    ws = wb["Q03"]
    assert ws.cell(1, 1).value.startswith("Q03")
    assert any(ws.cell(r, 1).value == "No run found for this question."
               for r in range(1, ws.max_row + 1))
    _write_index(wb, [ENTRY], [meta], "Expert review")
    assert wb["Review status"].cell(4, 5).value == "NOT RUN"
